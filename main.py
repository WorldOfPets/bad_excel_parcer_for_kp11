import openpyxl
from scheldue import Scheldue
wb = openpyxl.load_workbook('16.11.2023.xlsx')

ws = wb.active
skip_arr = ["1 пара", "2 пара", "3 пара", "4 пара", "5 пара"]
print(f'Total number of rows: {ws.max_row}. And total number of columb: {ws.max_column}')

data_group = []
for gr in range(1, ws.max_column - 1):
    val = ws.cell(row=6, column=gr).value
    if val is not None:
        data_group_id = f"{val} ! {gr}"
        data_group.append(data_group_id)



current_group = "Test group"
def get_group_name(index: int):
    for group in data_group:
        if "!" in str(group):
            group_arr = str(group).split("!")
            group_id = int(str(group_arr[1]).strip())
            group_name = group_arr[0]
            
            if group_id >= index:
                #print(f"{index}-{group_id} : {group_name}")
                return group_name

data_cabinet = []
data_teacher = []
data_lesson = []
data_para = []
new_data_group = []
current_para = "1 пара"
for i in range(1, ws.max_row - 1):
    for j in range(1, ws.max_column - 1):
        val = ws.cell(row=i, column=j).value
        
        if val is None or str(val).strip() == "":
            continue
        elif str(val).strip() in skip_arr:
            current_para = str(val).strip()
        else:
            current_group = get_group_name(j-1)
            if "/" in str(val):
                data_cell = str(val).split("/")
                data_lesson.append(data_cell[0])
                data_teacher.append(data_cell[1])
            if len(str(val)) == 3 or len(str(val)) == 4 or val == "Академия КП":
                data_cabinet.append(val)
                data_para.append(current_para)
                new_data_group.append(current_group)

data_classes = []
for all_i in range(len(data_cabinet)):
    obj = Scheldue(data_lesson[all_i], data_teacher[all_i], data_cabinet[all_i], data_para[all_i], new_data_group[all_i])
    new_obj = {
        "lesson" : f"{data_lesson[all_i]}",
        "teacher" : f"{data_teacher[all_i]}",
        "cabinet" : f"{data_cabinet[all_i]}",
        "para" : f"{data_para[all_i]}",
        "group" : f"{new_data_group[all_i]}"
    }
    super_new_obj = f"{data_lesson[all_i]}!{data_teacher[all_i]}!{data_cabinet[all_i]}!{data_para[all_i]}!{new_data_group[all_i]}"
    data_classes.append(super_new_obj)

# def clear_array_classes(data_classes):
#     current_obj:Scheldue = None
#     for obj in data_classes:
#         obj:Scheldue = obj
#         if current_obj is None:
#             current_obj = obj
#             continue
#         if current_obj.cabinet == obj.cabinet and current_obj.group == obj.group and current_obj.lesson == obj.lesson and current_obj.para == obj.para and current_obj.teacher == obj.teacher:
#             data_classes.remove(obj)
#         else:
#             current_obj = obj
#     return data_classes

data_classes = list(set(data_classes))
clear_data_classes = []
for user in data_classes:
    user = user.split("!")
    para = user[3]
    lesson = user[0]
    teacher = user[1]
    cabinet = user[2]
    group = user[4]
    obj = Scheldue(lesson, teacher, cabinet, para, group)
    clear_data_classes.append(obj)


main = input("""
1. Кто окупировал кабинет?
2. Какие у папы суриката пары?
3. Кто из папы сурикатов свободен?
4. Какие пары у группы сурикатов?
5. С кем поменяться ключами?
6. Какой кабинет свободен?
Действие: """)
if str(main) == "1":
    cab = input("Кабинет: ")
    for user in sorted(clear_data_classes, key=lambda user: user.para):
        user:Scheldue = user
        if cab == str(user.cabinet):
            print(str(user))
elif str(main) == "2":
    teacher = input("Папа сурикат: ")
    for user in sorted(clear_data_classes, key=lambda user: user.para):
        user:Scheldue = user
        if teacher in user.teacher:
            print(str(user))
elif str(main) == "3":
    para = input("""
    1 пара
    2 пара
    3 пара
    4 пара
    5 пара
    Цифра: """)
    all_teacher = list(set(data_teacher))
    for user in sorted(clear_data_classes, key=lambda user: user.teacher): 
        user:Scheldue = user
        if user.para == f"{para} пара":
            if user.teacher in all_teacher:
                all_teacher.remove(user.teacher)
    for teach in all_teacher:
        print(teach)
elif str(main) == "4":
    all_group = list(set(data_group))
    str_all_group = [f"{group}" for group in all_group]
    str_all_group += "Имя группы: "
    group_name = input(str_all_group)
    for user in sorted(clear_data_classes, key=lambda user: user.para): 
        user:Scheldue = user
        if group_name in user.group:
            print(str(user))
elif str(main) == "5":
    para = input("Номер пары: ")
    key_cab = input("Ключ который у вас: ")
    key_need = input("Ключ который нужен: ")
    need = ""
    if int(para) < 5:
        for user in sorted(clear_data_classes, key=lambda user: user.para): 
            if user.para == f"{int(para) + 1} пара" and user.cabinet == key_cab:
                print(f"Ваш ключ нужен папе сурикату {user.teacher}")
            if user.para == f"{int(para)} пара" and user.cabinet == key_need:
                need = f"Ключ который нужен вам у папы суриката {user.teacher}"
    if need == "":
        print("Скорее всего ключ на охране.")
    else:
        print(need)
elif str(main) == "6":
    para = input("""
    1 пара
    2 пара
    3 пара
    4 пара
    5 пара
    Цифра: """)
    all_cabinet = list(set(data_cabinet))
    all_str_cabinet = []
    for item in all_cabinet:
        all_str_cabinet.append(str(item))
        
    for user in sorted(clear_data_classes, key=lambda user: user.cabinet): 
        user:Scheldue = user
        if user.para == f"{para} пара":
            if user.cabinet in all_str_cabinet:
                all_str_cabinet.remove(user.cabinet)
    for cabi in all_str_cabinet:
        print(cabi)


    
        #print(f"{user.para}/{user.lesson}/{user.teacher} in {user.cabinet} with {user.group}")

